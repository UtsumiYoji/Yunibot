import openpyxl

#メンバーの登録
def add_member(msg):
    #メンバーリストにアクセス
    wb = openpyxl.open('./datasheet.xlsx')
    sh = wb['メンバーリスト']

    #重複idのチェック
    for i in range(30):
        if str(sh.cell(row=i+2, column=2).value)==str(msg[1]):
            r_msg = '既に登録されています'
            return r_msg
    
    #登録
    for i in range(30):
        if str(sh.cell(row=i+2, column=2).value)=='None':
            sh.cell(row=i+2, column=1).value = msg[2]
            sh.cell(row=i+2, column=2).value = msg[1]
            wb.save('./datasheet.xlsx')

            r_msg = '登録しました'
            return r_msg

    #登録失敗    
    r_msg = 'メンバーリストが一杯です'
    return r_msg

#メンバーの除外
def remove_member(msg):
    #メンバーリストにアクセス
    wb = openpyxl.open('./datasheet.xlsx')
    sh = wb['メンバーリスト']

    #除外
    for i in range(30):
        if str(sh.cell(row=i+2, column=2).value)==str(msg[1]):
            sh.cell(row=i+2, column=1).value = None
            sh.cell(row=i+2, column=2).value = None
            wb.save('./datasheet.xlsx')

            r_msg = '除外しました'
            return r_msg
    
    #除外失敗    
    r_msg = '除外対象が見つかりませんでした'
    return r_msg

#メンバーの確認
def all_member():
    wb = openpyxl.open('./datasheet.xlsx')
    sh = wb['メンバーリスト']

    name, memid, mension = '', '', ''
    for i in range(30):
        if not str(sh.cell(row=i+2, column=2).value)=='None':
            name += str(sh.cell(row=i+2, column=1).value) + '\n'
            memid += str(sh.cell(row=i+2, column=2).value) + '\n'
            mension += str('<@'+ sh.cell(row=i+2, column=2).value + '>') + '\n'
    
    return name, memid, mension