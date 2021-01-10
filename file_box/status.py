import datetime
import openpyxl

#残り予約の確認
def reserve():
    wb = openpyxl.open('./datasheet.xlsx', data_only=True)
    sh = wb['メンバーリスト']
    rev_list = [''] * 5

    #予約者の把握
    for y in range(2, 31):
        for x in range(3, 6):
            cell_data = str(sh.cell(row=y, column=x).value).split(',')
            #なんか予約があったら…
            if not cell_data == 'None':
                if cell_data[0] == '1':
                    rev_list[0] += '\n' + sh.cell(row=y, column=1).value + \
                        '(' + cell_data[1] + ')'

                elif cell_data[0] == '2':
                    rev_list[1] += '\n' + sh.cell(row=y, column=1).value + \
                        '(' + cell_data[1] + ')'
                
                elif cell_data[0] == '3':
                    rev_list[2] += '\n' + sh.cell(row=y, column=1).value + \
                        '(' + cell_data[1] + ')'
                
                elif cell_data[0] == '4':
                    rev_list[3] += '\n' + sh.cell(row=y, column=1).value + \
                        '(' + cell_data[1] + ')'
                
                elif cell_data[0] == '5':
                    rev_list[4] += '\n' + sh.cell(row=y, column=1).value + \
                        '(' + cell_data[1] + ')'
    
    #持越し者の把握
    for i in range(2, 31):
        co_num = sh.cell(row=i, column=7).value
        if co_num == 1:
            rev_list[0] += '\n' + sh.cell(row=i, column=1).value + '(持ち越し)'
        elif co_num == 2:
            rev_list[1] += '\n' + sh.cell(row=i, column=1).value + '(持ち越し)'
        elif co_num == 3:
            rev_list[2] += '\n' + sh.cell(row=i, column=1).value + '(持ち越し)'
        elif co_num == 4:
            rev_list[3] += '\n' + sh.cell(row=i, column=1).value + '(持ち越し)'
        elif co_num == 5:
            rev_list[4] += '\n' + sh.cell(row=i, column=1).value + '(持ち越し)'
    
    #空白の把握
    for i in range(5):
        if len(rev_list[i]) == 0:
            rev_list[i] += '予約なし'

    return rev_list

def carry_over():
    wb = openpyxl.open('./datasheet.xlsx')
    sh = wb['メンバーリスト']

    r_msg = ''

    for i in range(2, 31):
        co_num = str(sh.cell(row=i, column=7).value)

        #持越しが登録されてたら
        if not co_num == 'None':
            comm = sh.cell(row=i, column=8).value.split(',')
            r_msg += sh.cell(row=i, column=1).value+'('+co_num + comm[1] + comm[0] + ')\n' + comm[2] + '\n\n'
    
    #持越しが一件もない場合
    if not 'r_msg' in locals():
        r_msg = "持越しなし"
    
    return r_msg

def remaing():
    wb = openpyxl.open('./datasheet.xlsx')
    sh = wb['メンバーリスト']

    #残凸数の確認
    msg = ''
    rem = 0
    cao = 0
    for i in range(30):
        if sh.cell(row=i+2, column=6).value == 3:
            msg += str(sh.cell(row=i+2, column=1).value) + '@3'
            rem += 3
        
            if not sh.cell(row=i+2, column=7).value is None:
                msg += '+co' + str(sh.cell(row=i+2, column=7).value)
                cao += 1
            msg += '\n'

    for i in range(30):
        if sh.cell(row=i+2, column=6).value == 2:
            msg += str(sh.cell(row=i+2, column=1).value) + '@2'
            rem += 2

            if not sh.cell(row=i+2, column=7).value is None:
                msg += '+co' + str(sh.cell(row=i+2, column=7).value)
                cao += 1
            msg += '\n'
    
    for i in range(30):
        if sh.cell(row=i+2, column=6).value == 1:
            msg += str(sh.cell(row=i+2, column=1).value) + '@1'
            rem += 1
        
            if not sh.cell(row=i+2, column=7).value is None:
                msg += '+co' + str(sh.cell(row=i+2, column=7).value)
                cao += 1
            msg += '\n'
    
    for i in range(30):
        if sh.cell(row=i+2, column=6).value == 0:
            if not sh.cell(row=i+2, column=7).value is None:
                msg += str(sh.cell(row=i+2, column=1).value) + '@0'
                msg += '+co' + str(sh.cell(row=i+2, column=7).value)
                cao += 1
                msg += '\n'
    
    msg += '\n全体で' + str(rem) + '凸残っています'
    msg += '\n持越しが' + str(cao) + '件残っています'
    return msg, str(sh['J1'].value)

def endgame():
    #現在時刻をHH:MM形式で取得，55秒おき
    now = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=9)))
    now = now.strftime('%H:%M')

    #もしクローズできる時間じゃないなら帰る
    if not (now=='04:59'):
        return False

    #クローズ処理可能な時間ならExcelを開く
    wb = openpyxl.open('./datasheet.xlsx')
    sh = wb['メンバーリスト']

    #残凸数を記録しておく
    remain_sum = 0
    for i in range(30):
        remain_sum += sh.cell(row=i+2, column=6).value
    
    #各種数値をリセット
    for i in range(30):
        sh.cell(row=i+2, column=3).value = None
        sh.cell(row=i+2, column=4).value = None
        sh.cell(row=i+2, column=5).value = None
        sh.cell(row=i+2, column=6).value = 3
        sh.cell(row=i+2, column=7).value = None
        sh.cell(row=i+2, column=8).value = None

    wb.save('./datasheet.xlsx')

    #45凸以上あったらメッセージの送信
    if remain_sum > 45:
        return False
    
    #残凸のコマンドの呼び出し
    r_msg = remaing()[0]

    return r_msg
    