import openpyxl

def setin(msg, userid):
    #1~5の域を出た数値だったら怒る
    if not (0 < int(msg[1]) < 6):
        r_msg = '予約対象のボスは数値1~5で指定してください'
        return r_msg

    #コメントがない場合ダミーテキストを入れる
    if len(msg) == 2:
        msg.append('-')

    wb = openpyxl.open('./datasheet.xlsx')
    sh = wb['メンバーリスト']

    for i in range(30):
        if sh.cell(row=2+i, column=2).value == str(userid):
            #空いてる予約場所にいれていく
            if sh.cell(row=2+i, column=3).value == None:
                sh.cell(row=2+i, column=3).value = msg[1]+','+msg[2]

            elif sh.cell(row=2+i, column=4).value == None:
                sh.cell(row=2+i, column=4).value = msg[1]+','+msg[2]

            elif sh.cell(row=2+i, column=5).value == None:
                sh.cell(row=2+i, column=5).value = msg[1]+','+msg[2]
            else:
                r_msg = '3枠予約済みです．予約の取り消し後に再実行してください．'
                return r_msg

            #予約の完了
            wb.save('./datasheet.xlsx')
            r_msg = str(msg[1]) + 'ボスへの予約を受け付けました'
            return r_msg
    
    #メンバーが見つからない場合
    r_msg = '登録されていないメンバーです．リーダーに連絡してください．'
    return r_msg

def delete(msg, userid):
    #1~5の域を出た数値だったら怒る
    if not (0 < int(msg[1]) < 6):
        r_msg = '予約対象のボスは数値1~5で指定してください'
        return r_msg

    wb = openpyxl.open('./datasheet.xlsx')
    sh = wb['メンバーリスト']

    for i in range(30):
        if sh.cell(row=2+i, column=2).value == str(userid):
            #一致する予約場所を消す
            if str(sh.cell(row=2+i, column=3).value).split(',')[0] == msg[1]:
                sh.cell(row=2+i, column=3).value = None

            elif str(sh.cell(row=2+i, column=4).value).split(',')[0] == msg[1]:
                sh.cell(row=2+i, column=4).value = None

            elif str(sh.cell(row=2+i, column=5).value).split(',')[0] == msg[1]:
                sh.cell(row=2+i, column=5).value = None
            else:
                r_msg = str(msg[1]) + 'ボスへの予約が見つかりませんでした'
                return r_msg

            #予約の完了
            wb.save('./datasheet.xlsx')
            r_msg = str(msg[1]) + 'ボスへの予約を取り消しました'
            return r_msg
    
    #メンバーが見つからない場合
    r_msg = '登録されていないメンバーです．リーダーに連絡してください．'
    return r_msg

def carry_over(msg, userid):
    if not (int(msg[1]) < 6):
        r_msg = '持越しのボスは数値1~5で指定してください\n持越しの消化を宣言するときは0にして下さい'
        return r_msg
    
    wb = openpyxl.open('./datasheet.xlsx')
    sh = wb['メンバーリスト']

    for i in range(30):
        #idが一致する人の所に持ち越しを登録
        if sh.cell(row=2+i, column=2).value == str(userid):
            if int(msg[1]) == 0:
                sh.cell(row=2+i, column=7).value = None
                r_msg = '持越しを保持していない状態として登録しました'
            else:
                #情報が足りてるか
                if len(msg) < 4:
                    r_msg = "秒数，編成の情報がありません"
                    return r_msg

                #秒数の認識
                time = msg[2]
                #数字が入力されているか
                try:
                    time = int(time)
                    #有効な範囲内か
                    if not (20 < time < 91):
                        r_msg = "持越し秒数は21~90秒の間で入力して下さい．"
                        return r_msg
                except:
                    r_msg = '持越し秒数を認識出来ませんでした．数値のみで入力してください．'
                    return r_msg

                #編成の認識
                if "物理" in msg[3]:
                    team = "物理"
                elif "魔法" in msg[3]:
                    team = "魔法"
                elif "ニャル" in msg[3]:
                    team = "ニャル"
                else:
                    r_msg = "編成を認識出来ませんでした．\n物理・魔法・ニャルのいずれか入力してください．"
                    return r_msg
                
                #メッセージ無しの場合ダミーテキストを挿入
                if len(msg) == 4:
                    msg.append('-')
                
                #持越しボスの登録
                sh.cell(row=2+i, column=7).value = int(msg[1])

                #秒数，編成，コメントの書き込み
                sh.cell(row=2+i, column=8).value = msg[2] + 's,' + team + ',' + msg[4]

                r_msg = str(msg[1]) + 'ボスへ持越し登録を行いました．\n対象のボスが来たら通知します．'

            #予約の完了
            wb.save('./datasheet.xlsx')
            return r_msg

    #メンバーが見つからない場合
    r_msg = '登録されていないメンバーです．リーダーに連絡してください．'
    return r_msg

def atk(userid):
    wb = openpyxl.open('./datasheet.xlsx')
    sh = wb['メンバーリスト']

    for i in range(30):
        #idが一致する人の所の残凸を減らす，予約，持越し登録されてたら削除
        if sh.cell(row=2+i, column=2).value == str(userid):
            #持越し
            if not sh.cell(row=2+i, column=7).value is None:
                sh.cell(row=2+i, column=7).value = None
                sh.cell(row=2+i, column=8).value = None
                wb.save('./datasheet.xlsx')

                r_msg = '持越しの消費として処理しました'
                return r_msg, 0
            
            #3凸済み
            if sh.cell(row=2+i, column=6).value == 0:
                r_msg = '未消化の凸を確認できませんでした．\
                    \nタスキル等により修正が必要な場合は.fixをお使いください'
                return r_msg ,2
            
            #予約と一致して凸
            r_msg = '予約の消化として処理しました\n'
            if str(sh.cell(row=2+i, column=3).value).split(',')[0] == str(sh['J1'].value):
                sh.cell(row=2+i, column=3).value = None
                sh.cell(row=2+i, column=6).value -= 1

            elif str(sh.cell(row=2+i, column=4).value).split(',')[0] == str(sh['J1'].value):
                sh.cell(row=2+i, column=4).value = None
                sh.cell(row=2+i, column=6).value -= 1

            elif str(sh.cell(row=2+i, column=5).value).split(',')[0] == str(sh['J1'].value):
                sh.cell(row=2+i, column=5).value = None
                sh.cell(row=2+i, column=6).value -= 1
            
            #予約以外
            else:
                r_msg = ''
                sh.cell(row=2+i, column=6).value -= 1
            
            wb.save('./datasheet.xlsx')
            r_msg += '残り' + str(sh.cell(row=2+i, column=6).value) + '凸です'
            return r_msg, 1
        
    #メンバーが見つからない場合
    r_msg = '登録されていないメンバーです．リーダーに連絡してください．'
    return r_msg, 2

def fix(userid):
    wb = openpyxl.open('./datasheet.xlsx')
    sh = wb['メンバーリスト']

    for i in range(30):
        #idが一致する人の所の残凸を増やす
        if sh.cell(row=2+i, column=2).value == str(userid):
            #修正の必要がない場合
            if sh.cell(row=2+i, column=6).value == 3:
                r_msg = '1凸も記録されていません．\n修正処理は行いませんでした．'
                return r_msg

            #増やす処理
            sh.cell(row=2+i, column=6).value += 1
            wb.save('./datasheet.xlsx')
            r_msg = '修正処理を行いました．残り' + str(sh.cell(row=2+i, column=6).value) + '凸です．'
            return r_msg
        
    #メンバーが見つからない場合
    r_msg = '登録されていないメンバーです．リーダーに連絡してください．'
    return r_msg

def boss_change():
    wb = openpyxl.open('./datasheet.xlsx')
    sh = wb['メンバーリスト']

    #ボスの変更処理
    sh['J1'].value += 1
    if sh['J1'].value == 6:
        sh['J1'].value = 1
    wb.save('./datasheet.xlsx')
    
    msg = str(sh['J1'].value) + "ボスが来ました！\n予約者\n"
    #予約者の確認
    for i in range(30):
        #予約をリスト化
        for reserve_list_cell in sh['C'+str(i+2)+':E'+str(i+2)]:
            reserve_list = []
            for j in range(len(reserve_list_cell)):
                reserve_list.append(str(reserve_list_cell[j].value).split(',')[0])

        #予約リストの中に含まれていたら…
        if str(sh['J1'].value) in reserve_list:
            msg += '<@'+str(sh.cell(row=2+i, column=2).value)+'> '
    
    #持越し登録者の確認
    msg += '\n持越し登録者\n'
    for i in range(30):
        if sh.cell(row=i+2, column=7).value == sh['J1'].value:
            msg += '<@'+str(sh.cell(row=2+i, column=2).value)+'> '
    
    return msg

def boss_set(num):
    #受け取った値が正常か
    if not (0<num<6):
        r_msg = 'ボスの値は1~6を指定してください．'
    
    else:
        wb = openpyxl.open('./datasheet.xlsx')
        sh = wb['メンバーリスト']

        #ボスの値を変更
        sh['J1'].value = num
        wb.save('./datasheet.xlsx')

        r_msg = '現在のボスを' + str(num) + 'に変更しました'

    return r_msg
    